# processing.py
# Performs all the preproccessing required on plaintext for use with the SVM

from nltk import WordNetLemmatizer, pos_tag
from nltk.corpus import stopwords, wordnet
from sklearn.feature_extraction.text import TfidfVectorizer
from collections import Counter
import openpyxl as pyxl
import numpy as np
import pandas as pd
import math, os, re, pdb

ROOT = os.path.dirname(os.path.abspath(__file__))

def process(sentence):
	'''Processes the string 'sentence' in such a way that it can be used with the SVM'''

	# This whole thing? wrong for the moment.

	# Invalid sentence type
	if not isinstance(sentence, str):
		raise Exception("Processing Error: type must be str")
		return # unnecessary return but maintains my sanity

	# Initialize Lemmatizer
	lemmatizer = WordNetLemmatizer()

	# Convert to Lowercase
	sentence = sentence.lower()

	# Replace Nonletter Characters with Spaces
	sentence = re.sub('[^a-z]', " ", sentence)

	# Tokenize Sentence
	sentence = [word for word in sentence.split() if word != ' ' and len(word) > 1]

	# Remove Stop Words
	stop_words = set(stopwords.words("english"))
	sentence = [word for word in sentence if word not in stop_words]

	# Lemmatize
	sentence = [lemmatizer.lemmatize(word, get_wordnet_pos(word)) for word in sentence]

	# Calculate TF-IDF Score
	tfidf_vectorizer = TfidfVectorizer(lowercase=False, ngram_range=(1,2))

	vectors = tfidf_vectorizer.fit_transform(sentence)
	pdb.set_trace()
	return vectors


def preprocess():
	'''Performs all the preprocessing necessary to train the SVM (we don't want to train it every time we run the VA)'''
	# Initialize Lemmatizer
	lemmatizer = WordNetLemmatizer()

	# Load Data from intents.xlsx
	wb = pyxl.load_workbook('intents.xlsx')
	sheet = wb.active

	labels = []
	data = []

	# Iterate Col by Col
	for col in sheet.iter_cols(min_row=1, max_col=sheet.max_column, max_row=sheet.max_row, values_only=True):
		data.append([a for a in col[1:] if a])
		# pdb.set_trace()
		labels += [col[0]] * len(data[-1])

	for i, v in enumerate(data):
		row = []
		for j, k in enumerate(data[i]):
			# Convert to Lowercase
			sentence = k.lower()

			# Replace Nonletter Characters with Spaces
			sentence = re.sub('[^a-z]', " ", sentence)

			# Tokenize Sentence
			sentence = [word for word in sentence.split() if word != ' ' and len(word) > 1]

			# Remove Stop Words
			stop_words = set(stopwords.words("english"))
			sentence = [word for word in sentence if word not in stop_words]

			# Lemmatize
			sentence = [lemmatizer.lemmatize(word, get_wordnet_pos(word)) for word in sentence]

			# Save Processed Sentence
			row.append(' '.join(sentence))

		data[i] = row

	# Lower Dimensionality of Data (Easier to do this now than earlier)
	data = [str(a) for b in data for a in b]

	# Remove Duplicates from Data and their Corresponding Labels
	for i in range(len(data) - 1, -1, -1):
		if data[i] in data[:i]:
			data.pop(i)
			labels.pop(i)

	# Calculate TF-IDF Score (Need to change this to BoW later)
	vectorizer = TfidfVectorizer(lowercase=False, ngram_range=(1,2))
	vectors = vectorizer.fit_transform(data)
	pdb.set_trace()
	return vectors


def get_wordnet_pos(word):
	'''Maps the given word to its (simplified) part of speech for use in lemmatization'''
	tag = pos_tag([word])[0][1][0].upper()

	tag_dict = {"J": wordnet.ADJ,
				"N": wordnet.NOUN,
				"V": wordnet.VERB,
				"R": wordnet.ADV}

	return tag_dict.get(tag, wordnet.NOUN) # return simplified POS (defaults to NOUN)


def save_t():
	pass


def view_intents():
	'''Quality of Life Function that returns intents.xmls as a dictionary'''
	# Load Data from intents.xlsx
	wb = pyxl.load_workbook('intents.xlsx')
	sheet = wb.active

	info = {}
	
	# Iterate Col by Col
	for col in sheet.iter_cols(min_row=1, max_col=sheet.max_column, max_row=sheet.max_row, values_only=True):
		info.update({col[0]:[a for a in col[1:] if a]})

	return info


def train_svm(data, labels):
	'''Trains the SVM with the given dataset and labels.'''
	pass




def main():
	# temp = process("In my opinon, the newer model is better")
	preprocess()
	pdb.set_trace()



if __name__ == '__main__':
	main()











