# classifier.py
# Performs query processing and prediction.
from nltk import WordNetLemmatizer, pos_tag
from nltk.corpus import stopwords, wordnet
from sklearn.feature_extraction.text import CountVectorizer
from sklearn.metrics import confusion_matrix, ConfusionMatrixDisplay
from sklearn.svm import SVC
import matplotlib.pyplot as plt
import openpyxl as pyxl
import numpy as np
import pandas as pd
import pickle as pkl
import math, argparse, os, re, joblib, pdb

# Argument Parsing
parser = argparse.ArgumentParser()
parser.add_argument('-t', '--train', help='retrains/saves the SVM using intents.xlsx', action='store_true')
parser.add_argument('-a', '--accuracy', help="calculates the SVM's accuracy using testing.xlsx", action='store_true')

ROOT = os.path.dirname(os.path.abspath(__file__))

# Yes, global variables are gross. They're also convenient.
global clf
global feature_names

CONFIDENCE_THRESHOLD = 0.3


def predict(sentence, verbose=False):
	'''Processes and predicts the string "sentence"'''
	# Invalid sentence type
	if not isinstance(sentence, str):
		raise Exception("Processing Error: type must be str")
		return # unnecessary return but maintains my sanity

	# Initialize Lemmatizer
	lemmatizer = WordNetLemmatizer()

	# Convert to Lowercase
	sentence = sentence.lower()

	# Replace Nonletter (and non-math-operator) Characters with Spaces
	sentence = re.sub('[^a-z\+\-\*\/]', " ", sentence)

	# Tokenize Sentence
	sentence = [word for word in sentence.split() if word != ' ' and (len(word) > 1 or word in ['+','-','*','/'])]

	# Remove Stop Words
	stop_words = set(stopwords.words("english"))
	sentence = [word for word in sentence if word not in stop_words]

	# Lemmatize
	sentence = [lemmatizer.lemmatize(word, get_wordnet_pos(word)) for word in sentence]

	# Remove Unique Words (Present in Query but not in feature_names)
	sentence = [word for word in sentence if word in feature_names]

	# Calculate Individual BoW Score
	vectors = np.zeros(len(feature_names))

	for word in sentence:
		vectors[feature_names.index(word)] += 1

	# Predict Function
	prediction = clf.predict(vectors.reshape(1,-1))

	# Check for Empty Sentence (Invalid Query) or Low Confidence
	confidence = np.max(clf.predict_proba(vectors.reshape(1,-1)))

	if verbose:
		print(f"\tSentence: {sentence}")
		print(f"\tPrediction: {prediction}")
		print(f"\tConfidence: {confidence}")

	if not sentence or confidence < CONFIDENCE_THRESHOLD:
		return "unknown"

	# Return Prediction
	return str(prediction[0])


def preprocess():
	'''Performs all the necessary preprocessing, trains the SVM, and saves it'''
	# Initialize Lemmatizer
	lemmatizer = WordNetLemmatizer()

	# Load Data from intents.xlsx
	wb = pyxl.load_workbook('intents.xlsx')
	sheet = wb.active

	labels = []
	data = []

	# Iterate Col by Col
	for col in sheet.iter_cols(min_row=1, max_col=sheet.max_column, max_row=sheet.max_row, values_only=True):
		data.append([a for a in col[1:] if a])
		labels += [col[0]] * len(data[-1])

	for i, v in enumerate(data):
		row = []
		for j, k in enumerate(data[i]):
			# Convert to Lowercase
			sentence = str(k.lower())

			# Replace Nonletter (and non-math-operator) Characters with Spaces
			sentence = re.sub('[^a-z\+\-\*\/]', " ", sentence)

			# Tokenize Sentence
			sentence = [word for word in sentence.split() if word != ' ' and (len(word) > 1 or word in ['+','-','*','/'])]
			
			# Remove Stop Words
			stop_words = set(stopwords.words("english"))
			sentence = [word for word in sentence if word not in stop_words]

			# Lemmatize
			sentence = [lemmatizer.lemmatize(word, get_wordnet_pos(word)) for word in sentence]

			# Save Processed Sentence
			row.append(' '.join(sentence))

		data[i] = row

	# Lower Dimensionality of Data (Easier to do this now than earlier)
	data = [str(a) for b in data for a in b if a]

	# Sort Data/Labels Alphabetically ### unnecessary, just here for ease of reading
	data, labels = zip(*sorted(zip(data, labels)))

	# Calculate BoW Counts
	vectorizer = CountVectorizer(token_pattern="[a-z\*\/\+\-]+", ngram_range=(1,2))
	vectors = vectorizer.fit_transform(data)

	# Make DataFrame and Save ### unnecessary, just using this for sanity
	# frame = pd.DataFrame(vectors.toarray(), columns=vectorizer.get_feature_names())
	# frame.to_excel(os.path.join(ROOT, "bow_visualized.xlsx"))

	print("Training SVM...\r", end='')

	# Create SVM
	clf = SVC(kernel='linear', probability=True)
	clf.fit(vectors, labels)

	print("SVM successfully trained.")

	# Save SVM
	save_svm(clf)

	# Save Feature Names
	save_corpus([a for a in vectorizer.get_feature_names_out()])


def get_wordnet_pos(word):
	'''Maps the given word to its (simplified) part of speech for use in lemmatization'''
	tag = pos_tag([word])[0][1][0].upper()

	tag_dict = {"J": wordnet.ADJ,
				"N": wordnet.NOUN,
				"V": wordnet.VERB,
				"R": wordnet.ADV}

	return tag_dict.get(tag, wordnet.NOUN) # return simplified POS (defaults to NOUN)


def view_intents():
	'''Quality of Life Function that returns intents.xmls as a dictionary'''
	# Load Data from intents.xlsx
	wb = pyxl.load_workbook('intents.xlsx')
	sheet = wb.active

	info = {}
	
	# Iterate Col by Col
	for col in sheet.iter_cols(min_row=1, max_col=sheet.max_column, max_row=sheet.max_row, values_only=True):
		info.update({col[0]:[a for a in col[1:] if a]})

	return info


def save_svm(clf):
	'''Saves the trained SVM to trained_SVM.joblib for later use'''
	joblib.dump(clf, os.path.join(ROOT, "trained_SVM.joblib"))
	print("SVM successfully saved.") 


def load_svm_corpus():
	'''Loads the Pre-Trained SVM and feature names for use.'''
	global clf
	global feature_names

	# Load SVM
	clf = joblib.load(os.path.join(ROOT, "trained_SVM.joblib"))

	# Load Feature Names
	with open(os.path.join(ROOT, "feature_names.pkl"), "rb") as f:
		feature_names = pkl.load(f)


def save_corpus(feature_names):
	'''Saves the corpus of feature names into feature_names.pkl.
	This will only be used initially / if new intents are added.'''
	with open(os.path.join(ROOT, "feature_names.pkl"), "wb") as f:
		pkl.dump(feature_names, f)


def get_accuracy():
	'''Uses testing.xlsx to determine the model's accuracy.'''
	# Load Corpus
	load_svm_corpus()

	# Load Data from intents.xlsx
	wb = pyxl.load_workbook('testing.xlsx')
	sheet = wb.active

	labels = []
	data = []

	print("Reading testing dataset...\r", end='')
	# Iterate Col by Col
	for col in sheet.iter_cols(min_row=1, max_col=sheet.max_column, max_row=sheet.max_row, values_only=True):
		data.append([a for a in col[1:] if a])
		labels += [col[0]] * len(data[-1])

	# Flatten Data
	data = [a for b in data for a in b]

	print("Predicting queries...     \r", end='')
	# Get Prediction for Each Query in Data
	predictions = [predict(sentence) for sentence in data]

	print("Done.                     ")

	# Create Confusion Matrix
	matrix = confusion_matrix(labels, predictions)
	accuracy = np.divide(np.sum(matrix.diagonal()), np.sum(matrix))

	vmatrix = ConfusionMatrixDisplay.from_predictions(labels, predictions, xticks_rotation='vertical', cmap='Blues', colorbar=False)
	vmatrix.figure_.tight_layout()
	# plt.show()

	return f"Accuracy: {round(accuracy * 100, 3)}% ({np.sum(matrix.diagonal())}/{np.sum(matrix)})"





def main(args):
	# Retrain the SVM
	if args.train:
		preprocess()
		return

	# Get the SVM's Accuracy
	if args.accuracy:
		accuracy = get_accuracy()
		print(accuracy)
		return

	# This is just for testing purposes (and also to retrain the model if need be)
	load_svm_corpus()

	while True:
		query = input("\nEnter Query > ")

		if query == 'q':
			break

		prediction = predict(query, True)

		print(f"\tPrediction: {prediction}")


if __name__ == '__main__':
	main(parser.parse_args())






